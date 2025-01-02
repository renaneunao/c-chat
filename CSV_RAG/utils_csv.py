import pandas as pd
from pydantic import BaseModel
from io import BytesIO
from langchain.agents.agent_types import AgentType
from langchain_experimental.agents.agent_toolkits import create_pandas_dataframe_agent
from langchain_core.tools import StructuredTool
from dotenv import load_dotenv
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table

# noinspection PyUnresolvedReferences
from UTILS.utils import atualizar_historico, gerar_contexto

load_dotenv()

class DataFrameInputSchema(BaseModel):
    df_name: str = "combined_df"
    df_json: str

    class Config:
        arbitrary_types_allowed = True

def converse_com_csv(user_input, llm, st, context, prefix, show_max=5, formatar_padrao_credicaf=True):
    # Mark audio as generated
    st.session_state.audio_generated = False
    print(f'O dataframe que chegou aqui tem {len(st.session_state["combined_df"])} linhas')

    combined_df = st.session_state.get('combined_df', pd.DataFrame())

    def formatar_excel(df, transformar_em_tabela=False, estilizar=True, aplicar_bordas=True, aplicar_autosize=True):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Dados"

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        progresso_placeholder = st.empty()

        def atualizar_progresso(i, total, descricao):
            progresso = (i + 1) / total
            progresso_placeholder.progress(progresso, text=descricao)

        if estilizar:
            total_colunas = sheet.max_column
            for col in range(1, total_colunas + 1):
                cell = sheet.cell(row=1, column=col)
                cell.font = Font(color="00FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="003641", end_color="003641", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                atualizar_progresso(col - 1, total_colunas, "Estilizando cabeçalho")

        if aplicar_bordas:
            total_linhas = sheet.max_row
            for row in sheet.iter_rows(min_row=1, max_row=total_linhas, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                         top=Side(style='thin'), bottom=Side(style='thin'))
                atualizar_progresso(row[0].row - 1, total_linhas, "Aplicando bordas")

        if aplicar_autosize:
            col_list = list(sheet.columns)
            total_colunas = len(col_list)
            for idx, col in enumerate(col_list):
                max_length = 0
                column = col[0].column
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 5)
                column_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[column_letter].width = adjusted_width
                atualizar_progresso(idx, total_colunas, "Redimensionando colunas")

        if transformar_em_tabela:
            table = Table(displayName="Table1", ref=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")
            sheet.add_table(table)

        try:
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            if sheet.max_row <= 1:
                st.write("A planilha está vazia.")
            else:
                st.write("Planilha transformada e formatada:")

            st.download_button(
                label="Download Excel",
                data=excel_buffer,
                file_name="dados_formatados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Erro ao transformar o DataFrame em Excel: {e}")

    def transformar_dataframe_em_excel(df_name: str = "combined_df"):
        try:
            excel_buffer = BytesIO()
            combined_df.to_excel(excel_buffer, index=False)
            excel_buffer.seek(0)

            if combined_df.empty:
                st.write("O DataFrame está vazio.")
            else:
                st.write("DataFrame transformado em Excel:")
                st.dataframe(combined_df.head(show_max))

            if formatar_padrao_credicaf:
                formatar_excel(combined_df, transformar_em_tabela=True, estilizar=True, aplicar_bordas=True,
                               aplicar_autosize=True)
            else:
                st.download_button(
                    label="Download Excel",
                    data=excel_buffer,
                    file_name=f"{df_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"Erro ao transformar o DataFrame em Excel: {e}")

    def atualizar_dataframe(df: pd.DataFrame, df_name: str = "combined_df"):
        st.session_state[df_name] = df
        st.write("DataFrame atualizado:")
        st.dataframe(df)


    contexto = gerar_contexto(st, prefix=str(prefix) + str(context))

    tools = [
        StructuredTool(
            name='Transform_Dataframe_to_Excel',
            func=lambda df_name, df_json: transformar_dataframe_em_excel(df_name),
            description='Use this tool to transform the dataframe into an Excel file.',
            args_schema=DataFrameInputSchema
        ),
        StructuredTool(
            name='Save_And_Show_Dataframe',
            func=lambda df_name, df_json: atualizar_dataframe(pd.read_json(StringIO(df_json)), df_name),
            description='Use this tool to save and show the dataframe after any modification.',
            args_schema=DataFrameInputSchema
        ),
    ]

    agent = create_pandas_dataframe_agent(
        llm,
        combined_df,
        verbose=True,
        agent_type=AgentType.OPENAI_FUNCTIONS,
        allow_dangerous_code=True,
        prefix=contexto,
        extra_tools=tools,
        max_iterations=100,

    )

    resposta = agent.run(user_input)

    atualizar_historico(user_input, None, resposta, st)

    return resposta